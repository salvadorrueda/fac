import io
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from app.database import get_session
from app.models import Persona, Relacion, TipoRelacion

router = APIRouter(tags=["datos"])

COLS_PERSONA = ["id", "nombre", "primer_apellido", "segundo_apellido",
                "apodo", "fecha_nacimiento", "fecha_defuncion", "genero", "notas"]
COLS_RELACION = ["id", "persona_a_id", "persona_b_id", "tipo", "comentario"]


@router.get("/exportar")
def exportar(session: Session = Depends(get_session)):
    wb = Workbook()

    ws_p = wb.active
    ws_p.title = "Personas"
    ws_p.append(COLS_PERSONA)
    for cell in ws_p[1]:
        cell.font = Font(bold=True)
    for p in session.exec(select(Persona)).all():
        ws_p.append([p.id, p.nombre, p.primer_apellido, p.segundo_apellido,
                     p.apodo, p.fecha_nacimiento, p.fecha_defuncion, p.genero, p.notas])

    ws_r = wb.create_sheet("Relaciones")
    ws_r.append(COLS_RELACION)
    for cell in ws_r[1]:
        cell.font = Font(bold=True)
    for r in session.exec(select(Relacion)).all():
        ws_r.append([r.id, r.persona_a_id, r.persona_b_id, r.tipo.value, r.comentario])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=fac.xlsx"}
    )


@router.post("/importar")
def importar(file: UploadFile = File(...), session: Session = Depends(get_session)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El fichero debe ser .xlsx")

    wb = load_workbook(io.BytesIO(file.file.read()))

    if "Personas" not in wb.sheetnames or "Relaciones" not in wb.sheetnames:
        raise HTTPException(status_code=400,
                            detail="El fichero debe tener las hojas 'Personas' y 'Relaciones'")

    id_map: dict[int, int] = {}
    personas_creadas = 0

    ws_p = wb["Personas"]
    headers_p = [c.value for c in ws_p[1]]
    for row in ws_p.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        data = dict(zip(headers_p, row))
        old_id = data.pop("id", None)
        data = {k: str(v) if v is not None else None for k, v in data.items()}
        persona = Persona(**{k: v for k, v in data.items() if v is not None})
        session.add(persona)
        session.flush()
        if old_id is not None:
            id_map[int(old_id)] = persona.id
        personas_creadas += 1

    relaciones_creadas = 0
    ws_r = wb["Relaciones"]
    headers_r = [c.value for c in ws_r[1]]
    for row in ws_r.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        data = dict(zip(headers_r, row))
        data.pop("id", None)
        a_id = id_map.get(int(data["persona_a_id"]))
        b_id = id_map.get(int(data["persona_b_id"]))
        if a_id is None or b_id is None:
            continue
        relacion = Relacion(
            persona_a_id=a_id,
            persona_b_id=b_id,
            tipo=TipoRelacion(data["tipo"]),
            comentario=data.get("comentario"),
        )
        session.add(relacion)
        relaciones_creadas += 1

    session.commit()
    return {"personas_creadas": personas_creadas, "relaciones_creadas": relaciones_creadas}
